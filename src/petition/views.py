import io
from typing import Any, Dict, List, Optional

import openpyxl
import xlsxwriter
from pygooglechart import PieChart2D

from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Count, F, OuterRef, Q, Subquery, Sum
from django.http import (
    FileResponse,
    HttpResponse,
    HttpResponseForbidden,
    HttpResponseNotFound,
    HttpResponseRedirect,
)
from django.shortcuts import redirect, render, resolve_url
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views import generic
from django.views.generic.base import TemplateView, View
from django.views.generic.detail import DetailView
from django.views.generic.edit import CreateView, DeleteView, FormView

from account.models import User
from petition.models import Petition, PetitionNews, Vote

from .forms import ArchiveImportForm, PetitionCreateForm, PetitionNewsCreateForm

MAX_WORKSHEET_TITLE = 31


class Home(TemplateView):
    template_name = "home.html"

    def get_owned_petitions(self) -> List[Petition]:
        if not self.request.user.is_authenticated:
            return []

        return list(self.request.user.owned_petitions.all())

    def get_signed_petitions(self) -> List[Petition]:
        if not self.request.user.is_authenticated:
            return []

        # Query #1.2
        qs = self.request.user.signed_petitions.all()
        print("1.2:", qs.query)
        return list(qs)

    def get_trending_petitions(self) -> List[Petition]:
        limit = 20

        petitions = Petition.objects.filter(
            created_at__gt=timezone.now() - timezone.timedelta(days=14),
        ).order_by("-signatories_count")
        return list(petitions[:limit])

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx["signed_petitions"] = self.get_signed_petitions()
        ctx["trending_petitions"] = self.get_trending_petitions()
        ctx["owned_petitions"] = self.get_owned_petitions()
        return ctx


class Statistics(TemplateView):
    template_name = "statistics.html"


class Statistics0(TemplateView):
    template_name = "petition/list.html"

    def dispatch(self, request, *args, **kwargs):
        self.days = int(self.request.GET.get("days", "").strip())
        if not self.days:
            return HttpResponseRedirect(reverse("petition:statistics"))
        return super().dispatch(request, *args, **kwargs)

    def get_petitions(self) -> List[Petition]:
        if not self.request.user.is_authenticated:
            return []

        # Query #1.3
        petitions = Petition.objects.annotate(
            news_count=Subquery(
                PetitionNews.objects.filter(
                    created_at__gte=timezone.now() - timezone.timedelta(days=self.days),
                    petition_id=OuterRef("pk"),
                )
                .annotate(cnt=Count("id"))
                .values("cnt")[:1]
            )
        ).filter(news_count__gte=1)
        print("1.3: ", petitions.query)
        return list(petitions)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["list_title"] = "Petitions that had updates in the past {} days.".format(
            self.days
        )
        ctx["petitions"] = self.get_petitions()
        return ctx


class Statistics1(TemplateView):
    template_name = "petition/list.html"

    def dispatch(self, request, *args, **kwargs):
        self.votes = self.request.GET.get("votes", "").strip()
        if not self.votes:
            return HttpResponseRedirect(reverse("petition:statistics"))
        return super().dispatch(request, *args, **kwargs)

    def get_petitions(self) -> List[Petition]:
        if not self.request.user.is_authenticated:
            return []

        # Query #1.5
        petitions = Petition.objects.annotate(
            votes_count=Subquery(
                Vote.objects.filter(
                    created_at__gte=timezone.now() - timezone.timedelta(days=10),
                    petition_id=OuterRef("pk"),
                )
                .annotate(cnt=Count("id"))
                .values("cnt")[:1]
            )
        ).filter(votes_count__gte=self.votes)
        print("1.5: ", petitions.query)
        return list(petitions)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["list_title"] = (
            "Petitions which have collected >= {} votes in "
            "the past 10 days. votes:".format(self.votes)
        )
        ctx["petitions"] = self.get_petitions()
        return ctx


class Statistics2(TemplateView):
    template_name = "account/list.html"

    def get_users(self) -> List[Petition]:
        if not self.request.user.is_authenticated:
            return []

        # Query #2.1
        petitions = User.objects.raw(
            """
SELECT * FROM account_user u
WHERE array(SELECT v.petition_id FROM petition_vote v WHERE v.user_id = u.id) @>
     array(SELECT v.petition_id FROM petition_vote v WHERE v.user_id = %s) AND u.id != %s
            """,
            (self.request.user.pk, self.request.user.pk),
        )
        print("2.1: ", petitions.query)
        return list(petitions)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx[
            "list_title"
        ] = "All the users who voted for at least all the petitions I voted for"
        ctx["users"] = self.get_users()
        return ctx


class Statistics3(TemplateView):
    template_name = "account/list.html"

    def get_users(self) -> List[Petition]:
        if not self.request.user.is_authenticated:
            return []

        if self.request.user.owned_petitions.count() == 0:
            return []

        # Query #2.2
        users = User.objects.raw(
            """
SELECT * FROM account_user u
WHERE array(SELECT v.petition_id FROM petition_vote v WHERE v.user_id = u.id) @>
     array(SELECT p.id FROM petition_petition p WHERE p.author_id = %s) 
     AND u.id != %s
            """,
            (self.request.user.pk, self.request.user.pk),
        )
        print(users.query)
        return list(users)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["list_title"] = "Users who voted for all the petitions I've created"
        ctx["users"] = self.get_users()
        return ctx


class Statistics4(TemplateView):
    template_name = "account/list.html"

    def get_users(self) -> List[Petition]:
        if not self.request.user.is_authenticated:
            return []

        if self.request.user.owned_petitions.count() == 0:
            return []

        since = timezone.now() - timezone.timedelta(days=14)

        # Query #2.3
        users = User.objects.raw(
            """
SELECT * FROM account_user u
WHERE (SELECT count(1) FROM petition_petition p WHERE p.author_id = u.id AND created_at >= %s) >
     (SELECT count(1) FROM petition_petition p WHERE p.author_id = %s AND created_at >= %s) 
     AND u.id != %s
            """,
            (since, self.request.user.pk, since, self.request.user.pk),
        )
        print(users.query)
        return list(users)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx[
            "list_title"
        ] = "Users who have created more petitions in 2 weeks than I have"
        ctx["users"] = self.get_users()
        return ctx


class Search(TemplateView):
    template_name = "petition/search.html"

    def dispatch(self, request, *args, **kwargs):
        self.query = self.request.GET.get("q", "").strip()
        if not self.query:
            return HttpResponseRedirect(reverse("petition:archive"))
        return super().dispatch(request, *args, **kwargs)

    def get_petitions_by_title(self) -> List[Petition]:
        title_matched = Petition.objects.filter(title__icontains=self.query).order_by(
            "-created_at"
        )
        return list(title_matched)

    def get_petitions_by_description(self) -> List[Petition]:
        description_matched = (
            Petition.objects.filter(description__icontains=self.query)
            .exclude(title__icontains=self.query)
            .order_by("-created_at")
        )
        return list(description_matched)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx["search_query"] = self.query
        ctx["petitions_by_title"] = self.get_petitions_by_title()
        ctx["petitions_by_description"] = self.get_petitions_by_description()
        return ctx


class Archive(TemplateView):
    template_name = "petition/archive.html"

    def get_petitions(self) -> List[Petition]:
        petitions = Petition.objects.all().order_by("-created_at")
        return list(petitions)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx["petitions"] = self.get_petitions()
        return ctx


class ArchiveExport(View):
    def get_file(self) -> bytes:
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        petitions_ws = workbook.add_worksheet(name="Petitions List")

        petitions = Petition.objects.all().order_by("-created_at")

        petitions_ws.add_table(
            f"B2:F{2+len(petitions)}",
            {
                "columns": [
                    {"header": "Title"},
                    {"header": "Description"},
                    {"header": "Created At"},
                    {"header": "Author"},
                    {"header": "Signatories"},
                ]
            },
        )
        petitions_ws.set_column("B:C", 40)  # set width
        petitions_ws.set_column("D:F", 20)  # set width

        for idx, petition in enumerate(petitions, start=3):
            petitions_ws.write(f"B{idx}", petition.title)
            petitions_ws.write(f"C{idx}", petition.description)
            petitions_ws.write(f"D{idx}", petition.created_at.strftime("%d.%m.%Y"))
            petitions_ws.write(f"E{idx}", petition.author.full_name)
            petitions_ws.write(f"F{idx}", str(petition.signatories_count))

            signatories_ws = workbook.add_worksheet(
                name=petition.title[:MAX_WORKSHEET_TITLE]
            )

            votes = petition.votes.select_related("user").order_by("-created_at")
            signatories_ws.add_table(
                f"B2:C{2+len(votes)}",
                {
                    "columns": [
                        {"header": "Full name"},
                        {"header": "Sign date"},
                    ]
                },
            )
            signatories_ws.set_column("B:B", 30)  # set width
            signatories_ws.set_column("C:C", 15)  # set width

            for idx, vote in enumerate(votes, start=3):
                signatories_ws.write(f"B{idx}", vote.user.full_name)
                signatories_ws.write(f"C{idx}", vote.created_at.strftime("%d.%m.%Y"))
        workbook.close()

        buffer.seek(0)
        return buffer.read()

    def get(self, request, **kwargs):
        file = self.get_file()
        response = HttpResponse(
            file,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=archive.xlsx"
        return response


class ArchiveImport(FormView):
    form_class = ArchiveImportForm
    template_name = "petition/import.html"
    success_url = reverse_lazy("petition:archive")

    @staticmethod
    def get_user(full_name: str) -> Optional[User]:
        first_name, last_name = full_name.split(" ")
        user = User.objects.filter(first_name=first_name, last_name=last_name).first()
        return user

    @staticmethod
    def extract_petitions(workbook: openpyxl.workbook.Workbook) -> List[Dict[str, Any]]:
        petitions: List[Dict[str, Any]] = []
        for sheet in workbook.worksheets:
            if sheet.title == "Petitions List":
                for row in sheet.iter_rows(min_row=sheet.min_row + 1):
                    values = [i.value for i in row[1:]]
                    if None in values:
                        continue
                    title, description, created_at, author, signatories = values
                    petitions.append(
                        {
                            "title": title.replace("_x000D_", "\n"),
                            "description": description.replace("_x000D_", "\n"),
                            "created_at": timezone.datetime.strptime(
                                created_at, "%d.%m.%Y"
                            ),
                            "author": author.replace("_x000D_", "\n"),
                            "signatories": signatories,
                        }
                    )
                return petitions
        else:
            raise ValidationError('"Petitions List" sheet not found!')

    @staticmethod
    def extract_petition_signatories(sheet) -> List[Dict[str, Any]]:
        signatories = []
        for row in sheet.iter_rows(min_row=sheet.min_row + 1):
            values = [i.value for i in row[1:]]
            if None in values:
                continue
            full_name, created_at = values
            signatories.append(
                {
                    "full_name": full_name.replace("_x000D_", "\n"),
                    "created_at": timezone.datetime.strptime(created_at, "%d.%m.%Y"),
                }
            )
        return signatories

    def form_valid(self, form):
        file = form.cleaned_data["file"]
        workbook = openpyxl.load_workbook(file)
        with transaction.atomic():
            for petition_raw in self.extract_petitions(workbook):
                petition = Petition.objects.filter(title=petition_raw["title"]).first()
                if petition is None:
                    user = self.get_user(petition_raw["author"])
                    if user is None:
                        continue
                    petition = Petition.objects.create(
                        title=petition_raw["title"],
                        description=petition_raw["description"],
                        author_id=user.pk,
                    )
                elif petition.description != petition_raw["description"]:
                    # Edit a petition description if is has changed
                    petition.description = petition_raw["description"]
                    petition.save(update_fields=("description",))

                # Get a corresponding sheet with signatories
                try:
                    sheet_idx = workbook.sheetnames.index(
                        petition_raw["title"][:MAX_WORKSHEET_TITLE]
                    )
                except ValueError:
                    continue

                sheet = workbook.worksheets[sheet_idx]
                signatories = self.extract_petition_signatories(sheet)

                # Override all existing votes
                petition.votes.all().delete()
                for signatory in signatories:
                    user = self.get_user(signatory["full_name"])
                    if user is None:
                        continue

                    Vote.objects.create(
                        user_id=user.pk,
                        petition_id=petition.pk,
                        created_at=signatory["created_at"],
                    )
        return HttpResponseRedirect(self.success_url)


class ArchiveChart1Export(View):
    def get_file(self) -> bytes:
        colours = ["3374cd", "992220", "469b57", "e4e144", "cd3333", "749920"]
        since = timezone.now() - timezone.timedelta(days=14)
        # Query #1.1
        petitions = (
            Petition.objects.annotate(
                signatories_qty=Subquery(
                    Vote.objects.filter(
                        created_at__gte=since, petition_id=OuterRef("pk")
                    )
                    .annotate(qty=Count("id"))
                    .values("qty")[:1]
                )
            )
            .filter(signatories_qty__gte=1)
            .order_by("-signatories_count")
            .values("signatories_count", "title")[:5]
        )
        print("1.1: ", petitions.query)

        chart = PieChart2D(750, 400)
        chart.title = "5 most voted petitions since {}".format(
            since.strftime("%d.%m.%Y")
        )
        chart.add_data([i["signatories_count"] for i in petitions])
        chart.set_pie_labels(
            ["{} votes".format(i["signatories_count"]) for i in petitions]
        )
        chart.set_legend([i["title"] for i in petitions])
        chart.set_colours(colours[: len(petitions)])

        buffer = chart.download()
        return buffer

    def get(self, request, **kwargs):
        file = self.get_file()
        response = HttpResponse(
            file,
            content_type="image/png",
        )
        response["Content-Disposition"] = "attachment; filename=chart.png"
        return response


class ArchiveChart2Export(View):
    def get_file(self) -> bytes:
        colours = ["3374cd", "992220", "469b57", "e4e144", "cd3333", "749920"]
        since = timezone.now() - timezone.timedelta(days=14)

        petitions = (
            User.objects.all()
            .annotate(petitions_count=Count(F("owned_petitions")))
            .filter(petitions_count__gte=1)
            .order_by("-petitions_count")
            .values("first_name", "last_name", "petitions_count")[:5]
        )
        # Query #1.4
        print("1.4: ", petitions.query)

        chart = PieChart2D(750, 400)
        chart.title = "Top authors by a number of petitions".format(
            since.strftime("%d.%m.%Y")
        )
        chart.add_data([i["petitions_count"] for i in petitions])
        chart.set_pie_labels(
            ["{} petitions".format(i["petitions_count"]) for i in petitions]
        )
        chart.set_legend(
            ["{} {}".format(i["first_name"], i["last_name"]) for i in petitions]
        )
        chart.set_colours(colours[: len(petitions)])

        buffer = chart.download()
        return buffer

    def get(self, request, **kwargs):
        file = self.get_file()
        response = HttpResponse(
            file,
            content_type="image/png",
        )
        response["Content-Disposition"] = "attachment; filename=chart.png"
        return response


class PetitionCreate(LoginRequiredMixin, CreateView):
    model = Petition
    form_class = PetitionCreateForm
    template_name = "petition/create.html"

    def post(self, request, *args, **kwargs):
        self.object = None
        form = self.get_form()
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse("petition:detail", kwargs={"id": self.object.pk})

    def form_valid(self, form):
        form.instance.author_id = self.request.user.pk
        self.object = form.save()
        return HttpResponseRedirect(self.get_success_url())


class PetitionDetail(DetailView):
    model = Petition
    template_name = "petition/detail.html"
    pk_url_kwarg = "id"

    def get_is_owner(self):
        return self.request.user.pk == self.object.author_id

    def get_already_signed(self):
        return self.object.signatories.filter(pk=self.request.user.pk).exists()

    def get_latest_votes(self):
        return self.object.votes.order_by("-created_at")

    def get_news(self):
        return self.object.news.order_by("-created_at")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx["is_owner"] = self.get_is_owner()
        ctx["already_signed"] = self.get_already_signed()
        ctx["latest_votes"] = self.get_latest_votes()
        ctx["news"] = self.get_news()
        return ctx


class PetitionVote(LoginRequiredMixin, DetailView):
    model = Petition
    pk_url_kwarg = "id"

    def get(self, request, *args, **kwargs):
        petition = self.get_object()
        user = self.request.user
        Vote.objects.get_or_create(
            petition_id=petition.pk, user_id=user.pk, defaults={}
        )
        return redirect("petition:detail", id=petition.pk)


class PetitionNewsCreate(LoginRequiredMixin, CreateView):
    model = PetitionNews
    form_class = PetitionNewsCreateForm
    template_name = "petition_news/create.html"

    def dispatch(self, request, *args, **kwargs):
        self.petition = Petition.objects.filter(
            pk=self.kwargs["petition_id"],
            author_id=self.request.user.pk,
        ).first()
        if self.petition is None:
            return HttpResponseForbidden()

        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["petition"] = self.petition
        return ctx

    def post(self, request, *args, **kwargs):
        self.object = None
        form = self.get_form()
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse("petition:detail", kwargs={"id": self.object.petition.pk})

    def form_valid(self, form):
        form.instance.petition_id = self.petition.pk
        self.object = form.save()
        return HttpResponseRedirect(self.get_success_url())


class PetitionNewsDelete(LoginRequiredMixin, DeleteView):
    model = PetitionNews
    pk_url_kwarg = "id"

    def dispatch(self, request, *args, **kwargs):
        self.petition = Petition.objects.filter(
            news__id=self.kwargs["id"],
            author_id=self.request.user.pk,
        ).first()
        if self.petition is None:
            return HttpResponseForbidden()

        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return reverse("petition:detail", kwargs={"id": self.petition.pk})

    def get(self, request, *args, **kwargs):
        return self.delete(request, *args, **kwargs)

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()

        success_url = self.get_success_url()
        self.object.delete()
        return HttpResponseRedirect(success_url)

    def form_valid(self, form):
        form.instance.author_id = self.request.user.pk
        self.object = form.save()
        return HttpResponseRedirect(self.get_success_url())


class UserVotesList(TemplateView):
    template_name = "petition/list.html"

    def dispatch(self, request, *args, **kwargs):
        self.target_user = User.objects.filter(pk=self.kwargs["id"]).first()
        if self.target_user is None:
            return HttpResponseNotFound()
        return super().dispatch(request, *args, **kwargs)

    def get_petitions(self) -> List[Petition]:
        if not self.request.user.is_authenticated:
            return []

        petitions = Petition.objects.filter(votes__user_id=self.target_user.pk)

        return list(petitions)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["list_title"] = f"{self.target_user.full_name} votes:"
        ctx["petitions"] = self.get_petitions()
        return ctx
